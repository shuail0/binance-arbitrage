#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""策略管理工具 - 多服务器策略部署/启停/监控/账户余额/市价买入"""

import csv
import hashlib
import json
import os
import shlex
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from getpass import getpass
from pathlib import Path
from threading import Lock

import yaml
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ============================== 常量 ==============================

PROJECT_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = str(PROJECT_DIR / "keys" / "币安API.xlsx")
REMOTE_DIR = "/home/ubuntu/binance-py"
SSH_USER = "ubuntu"
CONDA_ENV = "binance"
DEPLOY_DEMO = os.getenv("BINANCE_DEMO", "0")
DEFAULT_SSH_PASSWORD = os.getenv("SSH_PASSWORD") or "LIshuai110110"

MAX_RETRIES = 3
RETRY_DELAY = 2
print_lock = Lock()


# ============================== 策略元数据 (单一数据源) ==============================

@dataclass(frozen=True)
class StrategySpec:
    mode: str          # screen 名后缀 (bnc-<mode>)
    name: str          # 目录名
    desc: str          # 中文描述
    config: str        # 相对项目根的配置文件路径
    hot_reload: bool   # main.py 是否监控 mtime
    # 通用字段名 → YAML 嵌套路径 (用于 view/edit/sync 通用化)
    fields: dict = field(default_factory=dict)
    # 快速编辑时展示的字段名
    quick_edit: tuple = ()


STRATEGIES = {
    "1": StrategySpec(
        mode="iv", name="instant_volume", desc="即时刷量",
        config="strategies/instant_volume/config.yaml",
        hot_reload=True,
        fields={
            "strategy_id": ("strategy_id",),
            "symbol": ("symbol",),
            "quantity": ("quantity",),
            "max_volume": ("max_volume",),
            "order_interval": ("order_interval",),
            "order_timeout": ("order_timeout",),
            "max_spread": ("max_spread",),
        },
        quick_edit=("quantity", "max_volume"),
    ),
    "2": StrategySpec(
        mode="spot", name="spot_volume_v2", desc="现货刷量V2",
        config="strategies/spot_volume_v2/config.yaml",
        hot_reload=True,
        fields={
            "strategy_id": ("strategy_id",),
            "symbol": ("symbol",),
            "quantity": ("quantity",),
            "max_volume": ("max_volume",),
            "order_interval": ("order_interval",),
            "step": ("step",),
            "num_levels": ("num_levels",),
            "max_net_buy": ("max_net_buy",),
        },
        quick_edit=("quantity", "max_volume"),
    ),
    "3": StrategySpec(
        mode="mm2", name="market_maker_v2", desc="做市策略V2",
        config="strategies/market_maker_v2/params.yaml",
        hot_reload=True,
        fields={
            "symbol": ("instrument_id",),
            "single_size": ("strategy", "single_size"),
            "step": ("strategy", "step"),
            "num_each_side": ("strategy", "num_of_order_each_side"),
            "max_net_buy": ("strategy", "maximum_net_buy"),
            "max_net_sell": ("strategy", "maximum_net_sell"),
            "max_volume": ("strategy", "max_volume"),
            "auto_close_on_exit": ("strategy", "auto_close_on_exit"),
        },
        quick_edit=("single_size", "max_volume"),
    ),
    "4": StrategySpec(
        mode="swap", name="swap_volume", desc="合约刷量",
        config="strategies/swap_volume/config.yaml",
        hot_reload=True,
        fields={
            "strategy_id": ("strategyId",),
            "symbol": ("symbol",),
            "leverage": ("leverage",),
            "margin_type": ("marginType",),
            "quantity": ("strategy", "quantity"),
            "max_volume": ("strategy", "maxVolume"),
            "order_interval": ("strategy", "orderInterval"),
            "step_size": ("strategy", "stepSize"),
            "num_each_side": ("strategy", "numOrdersEachSide"),
            "post_only": ("strategy", "postOnly"),
        },
        quick_edit=("quantity", "max_volume"),
    ),
}
BY_NAME = {s.name: s for s in STRATEGIES.values()}


# ============================== 通用工具 ==============================

def banner(title: str, ch: str = "=", width: int = 60):
    print(f"\n{ch * width}\n{title}\n{ch * width}")


def get_nested(data, path, default=None):
    if not path:
        return default
    cur = data
    for k in path:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def set_nested(data, path, value):
    cur = data
    for k in path[:-1]:
        cur = cur.setdefault(k, {})
    cur[path[-1]] = value


def coerce(s: str):
    """字符串 → 推断类型 (bool / int / float / str)"""
    low = s.lower()
    if low == "true":
        return True
    if low == "false":
        return False
    try:
        return int(s) if "." not in s and "e" not in low else float(s)
    except ValueError:
        return s


def load_yaml(path) -> dict:
    return yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}


def _safe_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _server_sort_key(sid):
    try:
        return (0, int(sid))
    except (ValueError, TypeError):
        return (1, str(sid))


# ============================== Excel 服务器列表 ==============================

def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _pick_col(header_map, *candidates):
    for c in candidates:
        k = _norm(c)
        if k in header_map:
            return header_map[k]
    for c in candidates:
        k = _norm(c)
        for hk, idx in header_map.items():
            if k in hk:
                return idx
    return None


def _norm_id(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_servers(excel_path: str) -> list:
    """从 Excel 读 (id, name, ip, api_key, key_path) 行, 过滤禁用"""
    DISABLED = {"否", "不可用", "停用", "no", "n", "false", "0"}
    wb = load_workbook(excel_path, data_only=True)
    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        hm = {_norm(t): i for i, t in enumerate(header_row) if t}
        ip_idx = _pick_col(hm, "IP", "服务器IP", "ip地址")
        if ip_idx is None:
            continue
        id_idx = _pick_col(hm, "序号", "编号", "id")
        name_idx = _pick_col(hm, "名称", "备注", "服务器名称")
        api_idx = _pick_col(hm, "API key", "api key", "apikey")
        key_idx = _pick_col(hm, "密钥路径", "密钥", "key path", "private key")
        enabled_idx = _pick_col(hm, "是否可用", "可用", "启用")

        def cell(row, idx):
            return str(row[idx] or "").strip() if idx is not None and idx < len(row) else ""

        servers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            ip = cell(row, ip_idx)
            if not ip:
                continue
            if enabled_idx is not None and enabled_idx < len(row) and _norm(row[enabled_idx]) in DISABLED:
                continue
            sid = _norm_id(row[id_idx] if id_idx is not None and id_idx < len(row) else len(servers) + 1)
            servers.append({"id": sid, "name": cell(row, name_idx),
                            "api_key": cell(row, api_idx), "key_path": cell(row, key_idx), "ip": ip})
        if servers:
            return servers
    return []


# ============================== SSH ==============================

def run_ssh(ip, password, command, timeout=30, retry=False):
    """通过 sshpass+bash 远程执行. retry: 仅传输错误重试, 应用层错误立即返回."""
    attempts = MAX_RETRIES if retry else 1
    for i in range(1, attempts + 1):
        cmd = [
            "sshpass", "-p", password, "ssh",
            "-o", "StrictHostKeyChecking=no", "-o", "ConnectTimeout=10",
            "-o", "ServerAliveInterval=5", f"{SSH_USER}@{ip}", "bash", "-s",
        ]
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, input=command)
            if r.returncode == 0 or i == attempts:
                return r.returncode == 0, r.stdout, r.stderr
        except subprocess.TimeoutExpired:
            if i == attempts:
                return False, "", f"SSH 超时 (>{timeout}s)"
        with print_lock:
            print(f"   ⚠️  第 {i} 次尝试失败, {RETRY_DELAY}s 后重试...")
        time.sleep(RETRY_DELAY)
    return False, "", "重试耗尽"


def upload_files(ip, password, local, remote, retry=True):
    """scp 上传 (本地→远端). 列表参数 + shell=False 防注入."""
    attempts = MAX_RETRIES if retry else 1
    last_err = ""
    for i in range(1, attempts + 1):
        cmd = [
            "sshpass", "-p", password, "scp",
            "-o", "StrictHostKeyChecking=no", "-o", "ConnectTimeout=10",
            "-r", str(local), f"{SSH_USER}@{ip}:{remote}",
        ]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if r.returncode == 0:
            return True, r.stderr
        last_err = r.stderr
        if i < attempts:
            time.sleep(RETRY_DELAY)
    return False, last_err


def upload_text(ip, password, content: str, remote_path: str, mode: int = 0o600) -> bool:
    """字符串写临时文件 → scp → chmod"""
    with tempfile.NamedTemporaryFile(mode="w", delete=False) as f:
        f.write(content)
        tmp = Path(f.name)
    try:
        ok, _ = upload_files(ip, password, tmp, remote_path, retry=True)
        if ok and mode:
            run_ssh(ip, password, f"chmod {mode:o} {shlex.quote(remote_path)}", timeout=5)
        return ok
    finally:
        tmp.unlink(missing_ok=True)


def make_env_content(api_key: str, key_path: str) -> str:
    key_basename = os.path.basename(key_path) if key_path else ""
    remote_key = f"{REMOTE_DIR}/{key_basename}" if key_basename else ""
    return (f"BINANCE_API_KEY={api_key}\n"
            f"BINANCE_PRIVATE_KEY_PATH={remote_key}\n"
            f"BINANCE_DEMO={DEPLOY_DEMO}\n")


def remote_python_cmd(script_path: str, args: list) -> str:
    """构造远程激活 conda 环境后执行 python 脚本的 bash 命令"""
    quoted = " ".join(shlex.quote(a) for a in args)
    return (f"cd {REMOTE_DIR} && "
            f". $HOME/miniconda3/etc/profile.d/conda.sh 2>/dev/null && "
            f"conda activate {CONDA_ENV} 2>/dev/null && "
            f"python3 {script_path} {quoted}")


# ============================== 交互 prompts ==============================

def select_servers(servers):
    print("\n可用服务器:")
    for s in servers:
        print(f"  {s['id']}. {s['ip']} ({s['name']})")
    inp = input("\n服务器编号 (逗号分隔 / 'all'): ").strip()
    if inp.lower() == "all":
        return servers
    ids = {x.strip() for x in inp.split(",") if x.strip()}
    return [s for s in servers if s["id"] in ids]


def select_strategy():
    print("\n可用策略:")
    for k, s in STRATEGIES.items():
        print(f"  {k}. {s.desc} ({s.name})")
    return STRATEGIES.get(input(f"选择策略 ({'/'.join(STRATEGIES)}): ").strip())


def get_password():
    if DEFAULT_SSH_PASSWORD:
        print("\n✅ 使用预设 SSH 密码")
        return DEFAULT_SSH_PASSWORD
    return getpass("\n请输入 SSH 密码: ")


# ============================== Remote Helper ==============================

def ensure_remote_helper(ip, password) -> bool:
    """本地 tools/remote_helper.py md5 与远端不一致则上传"""
    local = PROJECT_DIR / "tools" / "remote_helper.py"
    if not local.exists():
        return False
    remote = f"{REMOTE_DIR}/tools/remote_helper.py"
    local_md5 = hashlib.md5(local.read_bytes()).hexdigest()
    ok, out, _ = run_ssh(
        ip, password, f"md5sum {shlex.quote(remote)} 2>/dev/null | awk '{{print $1}}'",
        timeout=5, retry=False,
    )
    if ok and out.strip() == local_md5:
        return True
    run_ssh(ip, password, f"mkdir -p {REMOTE_DIR}/tools", timeout=5)
    ok, _ = upload_files(ip, password, local, remote, retry=True)
    return ok


def exec_remote_helper(ip, password, args: list, timeout=30) -> dict:
    """SSH 调用 remote_helper.py, 解析最后一行 JSON"""
    _, stdout, stderr = run_ssh(
        ip, password, remote_python_cmd("tools/remote_helper.py", args),
        timeout=timeout, retry=False,
    )
    lines = (stdout or "").strip().splitlines()
    for line in reversed(lines):
        if line.startswith("{"):
            try:
                return json.loads(line)
            except json.JSONDecodeError:
                continue
    return {"_transport_err": (stderr or stdout or "无输出")[:300]}


# ============================== Sync (.env + 策略配置) ==============================

def sync_server(server, password, specs_to_sync) -> bool:
    ip, sid, name = server["ip"], server["id"], server["name"]
    banner(f"同步服务器 {sid}: {ip} ({name})")

    # 1. .env
    print(f"[1] 同步 .env (API 密钥)...")
    if not upload_text(ip, password, make_env_content(server.get("api_key", ""), server.get("key_path", "")),
                       f"{REMOTE_DIR}/.env"):
        print(f"❌ .env 同步失败")
        return False
    print(f"✅ .env 已同步")

    # 上传密钥
    key_path = server.get("key_path", "")
    if key_path and os.path.exists(key_path):
        if upload_files(ip, password, key_path, f"{REMOTE_DIR}/")[0]:
            run_ssh(ip, password,
                    f"chmod 600 {REMOTE_DIR}/{shlex.quote(os.path.basename(key_path))}",
                    timeout=5)

    # 2. 策略配置
    valid = [s for s in specs_to_sync if (PROJECT_DIR / s.config).exists()]
    for missing in [s for s in specs_to_sync if not (PROJECT_DIR / s.config).exists()]:
        print(f"⚠️  跳过 {missing.desc}: 本地缺失")

    for i, spec in enumerate(valid, 2):
        print(f"[{i}] 同步 {spec.desc} 配置...")
        run_ssh(ip, password, f"mkdir -p {shlex.quote(REMOTE_DIR + '/' + os.path.dirname(spec.config))}",
                timeout=5)
        if not upload_files(ip, password, PROJECT_DIR / spec.config,
                            f"{REMOTE_DIR}/{spec.config}", retry=True)[0]:
            print(f"❌ 同步 {spec.desc} 失败")
            return False
        print(f"✅ {spec.desc} 已同步")

    print(f"✅ 服务器 {sid} 同步完成")
    return True


# ============================== Deploy ==============================

def deploy_server(server, password, full_setup=False, code_only=False):
    """部署单台服务器
    - code_only=True : 仅上传 .py/.sh, 不覆盖配置和 .env (excludes yaml/json/.env)
    - 默认           : 代码 + 配置 + .env + 密钥
    - full_setup=True: 默认基础上额外安装 miniconda + 创建 conda env + pip install
    """
    ip, sid, name = server["ip"], server["id"], server["name"]
    total = 5 if full_setup else (2 if code_only else 4)
    tag = " [仅代码]" if code_only else ""
    with print_lock:
        banner(f"[{sid}] 部署 {ip} ({name}){tag}")

    def step(i, msg):
        with print_lock:
            print(f"[{sid}] [{i}/{total}] {msg}")

    # 1. 远程目录
    step(1, "创建远程目录...")
    ok, _, err = run_ssh(ip, password,
                         f"mkdir -p {REMOTE_DIR}/logs {REMOTE_DIR}/strategies {REMOTE_DIR}/tools",
                         retry=True)
    if not ok:
        return False, server, f"mkdir 失败: {err}"

    # 2. rsync 代码
    step(2, "上传代码 (rsync)...")
    excl = ["--exclude=.env", "--exclude=__pycache__", "--exclude=logs/",
            "--exclude=.git", "--exclude=keys/"]
    if code_only:
        excl += ["--exclude=*.yaml", "--exclude=*.yml", "--exclude=*.json"]
    for sub in ("strategies", "tools"):
        local = PROJECT_DIR / sub
        if not local.exists():
            continue
        r = subprocess.run(
            ["sshpass", "-p", password, "rsync", "-az", *excl,
             "-e", "ssh -o StrictHostKeyChecking=no -o ConnectTimeout=10",
             f"{local}/", f"{SSH_USER}@{ip}:{REMOTE_DIR}/{sub}/"],
            capture_output=True, text=True, timeout=120,
        )
        if r.returncode != 0:
            return False, server, f"rsync {sub} 失败: {r.stderr[:200]}"
        with print_lock:
            print(f"[{sid}]    ✅ {sub} 已同步")

    for fname in ("run.sh", "requirements.txt"):
        local = PROJECT_DIR / fname
        if local.exists():
            upload_files(ip, password, local, f"{REMOTE_DIR}/{fname}")
    run_ssh(ip, password, f"chmod +x {REMOTE_DIR}/run.sh")

    if code_only:
        with print_lock:
            print(f"[{sid}] ✅ 部署成功（仅代码，配置未变）")
        return True, server, None

    # 3. .env + 密钥
    step(3, "上传 .env + 密钥...")
    if not upload_text(ip, password, make_env_content(server.get("api_key", ""), server.get("key_path", "")),
                       f"{REMOTE_DIR}/.env"):
        return False, server, "上传 .env 失败"
    key_path = server.get("key_path", "")
    if key_path and os.path.exists(key_path):
        upload_files(ip, password, key_path, f"{REMOTE_DIR}/")
        run_ssh(ip, password,
                f"chmod 600 {REMOTE_DIR}/{shlex.quote(os.path.basename(key_path))}", timeout=5)

    # 4. screen
    step(4, "安装 screen (若缺)...")
    run_ssh(ip, password,
            "command -v screen >/dev/null || sudo -n apt-get install -y screen 2>/dev/null || true",
            timeout=30)

    # 5. conda (可选)
    if full_setup:
        step(5, "安装 conda 环境...")
        run_ssh(ip, password, (
            "test -d $HOME/miniconda3 || ("
            "curl -fsSL https://repo.anaconda.com/miniconda/Miniconda3-latest-Linux-x86_64.sh -o /tmp/mc.sh && "
            "bash /tmp/mc.sh -b -p $HOME/miniconda3 && rm /tmp/mc.sh && "
            "$HOME/miniconda3/bin/conda init bash)"
        ), timeout=300, retry=True)
        run_ssh(ip, password, (
            f". $HOME/miniconda3/etc/profile.d/conda.sh && "
            f"(conda env list | grep -q '^{CONDA_ENV} ' || conda create -n {CONDA_ENV} python=3.12 -y) && "
            f"conda activate {CONDA_ENV} && pip install -r {REMOTE_DIR}/requirements.txt"
        ), timeout=300, retry=True)

    with print_lock:
        print(f"[{sid}] ✅ 部署成功")
    return True, server, None


# ============================== 单策略 action ==============================

def _action_banner(server, spec, verb: str):
    banner(f"服务器 {server['id']} ({server['ip']}) - {verb} {spec.desc}")


def start_strategy(server, spec, password):
    _action_banner(server, spec, "启动")
    ip, mode, name = server["ip"], spec.mode, spec.name

    # 检测已运行
    _, out, _ = run_ssh(ip, password, f'screen -list 2>/dev/null | grep "bnc-{mode}" || true')
    if f"bnc-{mode}" in out:
        _, pid, _ = run_ssh(ip, password, f'pgrep -f "python.*strategies/{name}/main.py" || true', timeout=5)
        if pid.strip():
            print(f"⚠️  策略已在运行 (screen: bnc-{mode})")
            return
        run_ssh(ip, password, f'screen -S "bnc-{mode}" -X quit 2>/dev/null', timeout=5)

    launch = f"""
if ! command -v screen >/dev/null; then
    echo "ERROR: screen 未安装, 先运行部署 (action 8) 或手动 'sudo apt-get install -y screen'"
    exit 1
fi
screen -dmS "bnc-{mode}" bash -c '
    . $HOME/miniconda3/etc/profile.d/conda.sh 2>/dev/null
    conda activate {CONDA_ENV} 2>/dev/null
    cd {REMOTE_DIR}
    python strategies/{name}/main.py
'"""
    run_ssh(ip, password, launch, timeout=10)
    time.sleep(2)
    _, out, _ = run_ssh(ip, password, f'screen -list 2>/dev/null | grep "bnc-{mode}" && echo OK', timeout=5)
    print(f"✅ 已启动 (screen: bnc-{mode})" if "OK" in out else "❌ 启动失败")


def stop_strategy(server, spec, password):
    _action_banner(server, spec, "停止")
    stop_cmd = f"""
PID=""
for p in $(pgrep -f "strategies/{spec.name}/main.py" 2>/dev/null); do
    case "$(cat /proc/$p/comm 2>/dev/null)" in python*) PID=$p; break;; esac
done
if [ -n "$PID" ]; then
    echo "发送 SIGINT 到 $PID..."
    kill -INT $PID
    for i in $(seq 1 30); do kill -0 $PID 2>/dev/null || break; sleep 1; done
    kill -0 $PID 2>/dev/null && {{ echo "强制 kill"; kill -9 $PID 2>/dev/null; }}
    echo "STOPPED"
else
    echo "NOT_RUNNING"
fi
screen -S "bnc-{spec.mode}" -X quit 2>/dev/null
"""
    _, out, err = run_ssh(server["ip"], password, stop_cmd, timeout=60)
    if "NOT_RUNNING" in out:
        print("⚠️  策略未运行")
    elif "STOPPED" in out:
        print("✅ 策略已停止")
    else:
        print(f"❌ 停止失败: {err.strip()}")


def _render_progress(volume: float, max_volume: float):
    if not max_volume or max_volume <= 0:
        return
    pct = min(volume / max_volume * 100, 100.0)
    bar_w = 40
    filled = int(bar_w * pct / 100)
    print(f"\n📍 进度: {volume:,.2f} / {max_volume:,.2f} USDT ({pct:.1f}%)")
    print(f"[{'█' * filled}{'░' * (bar_w - filled)}] {pct:.1f}%")


def _print_symbol_stats(stats: dict, symbol: str, max_volume: float):
    buy = _safe_float(stats.get("global_buy_volume"))
    sell = _safe_float(stats.get("global_sell_volume"))
    pnl = _safe_float(stats.get("global_pnl"))
    volume = buy + sell
    banner(f"📊 全局统计数据" + (f"（{symbol}）" if symbol else ""))
    print(f"💰 累计总成交额: {volume:,.2f} USDT")
    print(f"📈 买入成交额:   {buy:,.2f} USDT")
    print(f"📉 卖出成交额:   {sell:,.2f} USDT")
    tag = "盈利" if pnl > 0 else ("亏损" if pnl < 0 else "持平")
    print(f"💎 累计盈亏:     {pnl:+,.6f} USDT ({tag})")
    print(f"🕐 最后更新:     {stats.get('updated_at', 'N/A')}")
    _render_progress(volume, max_volume)


def view_status(server, spec, password):
    _action_banner(server, spec, "状态")
    ip = server["ip"]

    _, out, _ = run_ssh(ip, password, f'ps aux | grep "python.*strategies/{spec.name}/main.py" | grep -v grep')
    print("✅ 策略正在运行：" if out.strip() else "❌ 策略未运行")
    if out.strip():
        print(out.rstrip())

    _, out, _ = run_ssh(ip, password, f'screen -list 2>/dev/null | grep "bnc-{spec.mode}" || echo "无 screen 会话"')
    print(f"\nScreen: {out.strip()}")

    # config + state
    _, cfg_out, _ = run_ssh(ip, password, f"cat {shlex.quote(REMOTE_DIR + '/' + spec.config)} 2>/dev/null")
    config = {}
    max_volume = 0.0
    if cfg_out.strip():
        try:
            config = yaml.safe_load(cfg_out) or {}
            max_volume = _safe_float(get_nested(config, spec.fields.get("max_volume"), 0))
        except Exception:
            pass

    _, state_out, _ = run_ssh(ip, password, f"cat {REMOTE_DIR}/volume_state.json 2>/dev/null")
    if not state_out.strip():
        print(f"\n📊 volume_state.json 不存在（首次运行或未产生交易）")
        return
    try:
        all_state = json.loads(state_out)
    except json.JSONDecodeError as e:
        print(f"\n⚠️  解析 volume_state.json 失败: {e}")
        return

    symbol = get_nested(config, spec.fields.get("symbol"), "")
    stats = all_state.get(symbol, {}) if symbol else {}
    if stats:
        _print_symbol_stats(stats, symbol, max_volume)
    else:
        print(f"\n📊 未找到 {symbol or '该策略'} 的统计数据")


def view_logs(server, spec, password):
    _action_banner(server, spec, "日志 (最后 50 行)")
    ip = server["ip"]
    for pattern in (f"*{spec.mode}*.log", "*.log"):
        _, log_file, _ = run_ssh(ip, password, f'ls -t {REMOTE_DIR}/logs/{pattern} 2>/dev/null | head -1')
        log_file = log_file.strip()
        if log_file:
            break
    if not log_file:
        print("⚠️  日志文件不存在")
        return
    _, out, _ = run_ssh(ip, password, f"tail -50 {shlex.quote(log_file)}")
    if out.strip():
        print(f"📄 {log_file}")
        print(out)
    else:
        print("⚠️  日志为空")


def view_config(server, spec, password, display_mode="formatted"):
    _action_banner(server, spec, "配置")
    ip = server["ip"]
    config_path = f"{REMOTE_DIR}/{spec.config}"
    _, out, _ = run_ssh(ip, password, f"cat {shlex.quote(config_path)} 2>/dev/null || echo FILE_NOT_FOUND")
    if "FILE_NOT_FOUND" in out:
        print(f"❌ 配置文件不存在: {config_path}")
        return

    if display_mode == "raw":
        print("\n原始 YAML:")
        print(out)
    else:
        try:
            config = yaml.safe_load(out) or {}
            print(f"\n📋 {spec.desc} ({spec.name})  热加载: {'✅' if spec.hot_reload else '❌'}")
            print("─" * 60)
            for fname, path in spec.fields.items():
                val = get_nested(config, path, "N/A")
                print(f"  {fname:<20} = {val}")
        except Exception as e:
            print(f"❌ YAML 解析失败: {e}")
            print(out)

    # .env 脱敏
    print("\n📋 .env:")
    _, env_out, _ = run_ssh(ip, password, f"cat {REMOTE_DIR}/.env 2>/dev/null || echo FILE_NOT_FOUND")
    if "FILE_NOT_FOUND" in env_out:
        print("  .env 不存在")
        return
    for line in env_out.splitlines():
        if "=" not in line or line.startswith("#"):
            continue
        key, val = line.split("=", 1)
        key = key.strip()
        if key == "BINANCE_API_KEY" and len(val) > 10:
            print(f"  {key}={val[:6]}...{val[-4:]}")
        elif key == "BINANCE_PRIVATE_KEY_PATH":
            print(f"  {key}={os.path.basename(val)}")
        else:
            print(f"  {line}")


def update_config(server, spec, password):
    _action_banner(server, spec, "修改配置")
    ip = server["ip"]
    config_path = f"{REMOTE_DIR}/{spec.config}"
    ok, out, err = run_ssh(ip, password, f"cat {shlex.quote(config_path)}")
    if not ok:
        print(f"❌ 读取配置失败: {err}")
        return
    try:
        config = yaml.safe_load(out) or {}
    except Exception as e:
        print(f"❌ YAML 解析失败: {e}")
        return

    print("\n当前配置:")
    print(yaml.dump(config, allow_unicode=True, default_flow_style=False))

    print(f"修改模式:\n  1. 快速修改 ({', '.join(spec.quick_edit)})\n  2. 高级修改 (任意字段)")
    choice = input("(1/2): ").strip()
    updates = {}

    if choice == "1":
        for fname in spec.quick_edit:
            path = spec.fields[fname]
            cur = get_nested(config, path)
            v = input(f"  {fname} [当前: {cur}]: ").strip()
            if v:
                set_nested(config, path, coerce(v))
                updates[fname] = v
    elif choice == "2":
        print("\nkey=value 每行一个, 空行结束. 嵌套用点: strategy.quantity=500")
        while True:
            line = input("> ").strip()
            if not line:
                break
            if "=" not in line:
                print("  ⚠️  格式: key=value")
                continue
            key, val = line.split("=", 1)
            set_nested(config, key.strip().split("."), coerce(val.strip()))
            updates[key.strip()] = val.strip()

    if not updates:
        print("\n⚠️  未修改任何参数")
        return

    yaml_str = yaml.dump(config, allow_unicode=True, default_flow_style=False, sort_keys=False)
    print("\n上传新配置...")
    if upload_text(ip, password, yaml_str, config_path, mode=0):
        print("✅ 配置已更新")
        for k, v in updates.items():
            print(f"  {k}: {v}")
        print("\n💡 " + ("策略支持热加载, 无需重启" if spec.hot_reload else "需重启才生效"))
    else:
        print(f"❌ 上传失败")


# ============================== 账户余额 ==============================

ACCOUNT_LABELS = {"funding": "资金账户", "spot": "现货账户", "futures": "合约账户"}


def query_single_balance(server, password, accounts, assets):
    if not ensure_remote_helper(server["ip"], password):
        return None
    args = ["balance", "--accounts", accounts]
    if assets:
        args += ["--assets", assets]
    return exec_remote_helper(server["ip"], password, args, timeout=30)



def parallel_query_balance(servers, password, accounts, assets):
    print(f"\n🔍 并行查询 {len(servers)} 服务器...")
    results = []
    with ThreadPoolExecutor(max_workers=min(10, len(servers))) as ex:
        futures = {ex.submit(query_single_balance, s, password, accounts, assets): s for s in servers}
        for fut in as_completed(futures):
            srv = futures[fut]
            try:
                data = fut.result()
            except Exception as e:
                with print_lock:
                    print(f"[{srv['id']}] ❌ 异常: {e}")
                continue
            if data and not data.get("_transport_err"):
                results.append({"server": srv, "data": data})
                with print_lock:
                    print(f"[{srv['id']}] ✅ 完成")
            else:
                err = (data or {}).get("_transport_err", "未知错误")
                with print_lock:
                    print(f"[{srv['id']}] ❌ {err[:120]}")
    return results


def _collect_account_assets(results, acct):
    """聚合账户内所有非零资产 + 每服务器映射. USDT 优先, 其余字母序"""
    asset_set = set()
    per_server = {}
    for item in results:
        sid = item["server"]["id"]
        section = (item["data"] or {}).get(acct) or {}
        if section.get("err"):
            per_server[sid] = {"_err": section["err"]}
            continue
        m = {}
        for it in (section.get("items") or []):
            a = it.get("asset", "")
            if a:
                m[a] = (_safe_float(it.get("free")), _safe_float(it.get("locked")))
                asset_set.add(a)
        per_server[sid] = m
    return sorted(asset_set, key=lambda a: (0, a) if a == "USDT" else (1, a)), per_server


def _balance_blocks(results, accounts):
    """yields (title, assets, per_server_rows, totals)"""
    for acct in accounts:
        assets, per_server = _collect_account_assets(results, acct)
        rows = []
        totals = {a: [0.0, 0.0] for a in assets}
        for item in results:
            sid, name = item["server"]["id"], item["server"]["name"]
            m = per_server.get(sid, {})
            if "_err" in m:
                rows.append({"sid": sid, "name": name, "err": m["_err"]})
                continue
            row = {"sid": sid, "name": name, "cells": []}
            for a in assets:
                free, locked = m.get(a, (0.0, 0.0))
                row["cells"].append((free, locked))
                totals[a][0] += free
                totals[a][1] += locked
            rows.append(row)
        yield ACCOUNT_LABELS.get(acct, acct), assets, rows, totals


def print_balance_table(results, accounts):
    if not results:
        print("\n⚠️  无数据")
        return
    results.sort(key=lambda r: _server_sort_key(r["server"]["id"]))
    account_list = [a.strip() for a in accounts.split(",") if a.strip()]
    W_ID, W_NAME, W_NUM = 6, 14, 16
    fmt = lambda v: f"{v:>{W_NUM},.4f}"

    print(f"\n查询时间: {datetime.now():%Y-%m-%d %H:%M:%S} | 服务器数: {len(results)}")
    for title, assets, rows, totals in _balance_blocks(results, account_list):
        width = max(50, W_ID + 1 + W_NAME + len(assets) * 2 * (1 + W_NUM))
        print("\n" + "=" * width)
        print(title)
        print("=" * width)
        if not assets:
            errs = [r for r in rows if "err" in r]
            if errs:
                for r in errs:
                    print(f"  ❌ 服务器 {r['sid']}: {r['err'][:120]}")
            else:
                print("  (无非零余额)")
            continue
        header = f"{'服务器':<{W_ID}} {'名称':<{W_NAME}}"
        for a in assets:
            header += f" {a + '(可用)':>{W_NUM}} {a + '(冻结)':>{W_NUM}}"
        print(header)
        print("-" * width)
        for r in rows:
            if "err" in r:
                print(f"{r['sid']:<{W_ID}} {r['name']:<{W_NAME}}   ❌ {r['err'][:80]}")
                continue
            line = f"{r['sid']:<{W_ID}} {r['name']:<{W_NAME}}"
            for free, locked in r["cells"]:
                line += f" {fmt(free)} {fmt(locked)}"
            print(line)
        print("-" * width)
        total_row = f"{'合计':<{W_ID}} {'':<{W_NAME}}"
        for a in assets:
            tf, tl = totals[a]
            total_row += f" {fmt(tf)} {fmt(tl)}"
        print(total_row)
    print()


def export_balance_csv(results, accounts):
    if not results:
        print("⚠️  无数据可导出")
        return
    results.sort(key=lambda r: _server_sort_key(r["server"]["id"]))
    account_list = [a.strip() for a in accounts.split(",") if a.strip()]
    logs = PROJECT_DIR / "logs"
    logs.mkdir(exist_ok=True)
    filepath = logs / f"balance_{datetime.now():%Y%m%d_%H%M%S}.csv"

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["查询时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "服务器数", len(results)])
        w.writerow([])
        for title, assets, rows, totals in _balance_blocks(results, account_list):
            w.writerow([title])
            header = ["服务器", "名称"]
            for a in assets:
                header += [f"{a}(可用)", f"{a}(冻结)"]
            w.writerow(header)
            for r in rows:
                if "err" in r:
                    w.writerow([r["sid"], r["name"], f"错误: {r['err']}"])
                    continue
                row = [r["sid"], r["name"]]
                for free, locked in r["cells"]:
                    row += [f"{free:.4f}", f"{locked:.4f}"]
                w.writerow(row)
            total_row = ["合计", ""]
            for a in assets:
                tf, tl = totals[a]
                total_row += [f"{tf:.4f}", f"{tl:.4f}"]
            w.writerow(total_row)
            w.writerow([])
    print(f"\n✅ CSV 已导出: {filepath}")


def select_accounts():
    print("\n查询账户:")
    print("  1. 全部 (资金+现货+合约) [默认]")
    print("  2. 仅资金账户")
    print("  3. 仅现货账户")
    print("  4. 仅合约账户")
    print("  5. 自定义 (逗号分隔: funding,spot,futures)")
    choice = input("(1-5, 默认 1): ").strip() or "1"
    m = {"1": "funding,spot,futures", "2": "funding", "3": "spot", "4": "futures"}
    if choice in m:
        return m[choice]
    if choice == "5":
        custom = input("账户: ").strip().lower()
        valid = [a.strip() for a in custom.split(",")
                 if a.strip() in ACCOUNT_LABELS]
        return ",".join(valid) if valid else "funding,spot,futures"
    return "funding,spot,futures"


def query_balance_main(selected_servers):
    accounts = select_accounts()
    assets_input = input("\n资产筛选 (逗号分隔, 留空=所有非零, 例: USDT,BNB,BTC): ").strip()
    assets = ",".join(a.strip().upper() for a in assets_input.split(",") if a.strip())
    password = get_password()

    results = parallel_query_balance(selected_servers, password, accounts, assets)
    if not results:
        print("\n❌ 无结果")
        return
    print_balance_table(results, accounts)
    if input("\n导出 CSV? (y/N): ").strip().lower() == "y":
        export_balance_csv(results, accounts)


# ============================== 市价买入 ==============================

def buy_single_server(server, password, symbol, mode, amount):
    if not ensure_remote_helper(server["ip"], password):
        return None
    args = ["buy", "--symbol", symbol]
    args += ["--quote", amount] if mode == "usdt" else ["--quantity", amount]
    return exec_remote_helper(server["ip"], password, args, timeout=30)


def buybnb_main(selected_servers):
    symbol = (input("\n交易对 (默认 BNBUSDT): ").strip() or "BNBUSDT").upper()
    print("\n买入模式:\n  1. 按基础币数量 (quantity)\n  2. 按 USDT 金额 (quoteOrderQty)")
    mode = "usdt" if input("(1/2, 默认 1): ").strip() == "2" else "quantity"
    base = symbol[:3] or "BASE"
    prompt = "USDT 金额 (如 100)" if mode == "usdt" else f"{base} 数量 (如 0.5)"
    amount = input(f"\n每服务器 {prompt}: ").strip()
    try:
        if float(amount) <= 0:
            raise ValueError
    except ValueError:
        print(f"❌ 无效数字: {amount}")
        return

    label = f"{amount} USDT 等值 {base}" if mode == "usdt" else f"{amount} {base}"
    print(f"\n⚠️  即将在 {len(selected_servers)} 服务器市价买入 {label}")
    if input("确认 (yes/N): ").strip().lower() not in ("y", "yes"):
        print("已取消")
        return
    password = get_password()

    print("\n🚀 并行执行...")
    results = []
    with ThreadPoolExecutor(max_workers=min(10, len(selected_servers))) as ex:
        futures = {ex.submit(buy_single_server, s, password, symbol, mode, amount): s for s in selected_servers}
        for fut in as_completed(futures):
            srv = futures[fut]
            try:
                data = fut.result()
                if data:
                    results.append({"server": srv, "data": data})
            except Exception as e:
                with print_lock:
                    print(f"[{srv['id']}] ❌ 异常: {e}")

    results.sort(key=lambda r: _server_sort_key(r["server"]["id"]))
    sep = "=" * 100
    print(f"\n{sep}")
    print(f"{'服务器':<18}{'状态':<8}{'OrderID':>14}{'成交量':>14}{'成交额':>18}{'均价':>14}{'手续费':>22}")
    print("-" * 100)
    total_qty = total_quote = total_fee = 0.0
    ok = 0
    for r in results:
        srv, d = r["server"], r["data"]
        tag = f"{srv['id']}/{(srv['name'] or '')[:8]}"
        if d.get("_transport_err") or not d.get("ok"):
            err = (d.get("err") or d.get("_transport_err") or "失败")[:20]
            print(f"{tag:<18}{'失败':<8}{'-':>14}{'-':>14}{'-':>18}{'-':>14}{err:>22}")
            continue
        ok += 1
        qty = _safe_float(d.get("executedQty"))
        quote = _safe_float(d.get("cumQuoteQty"))
        fee_bnb = d.get("feeBNB") or "0"
        fee_other = d.get("feeOther") or ""
        total_qty += qty
        total_quote += quote
        total_fee += _safe_float(fee_bnb)
        fee_str = f"{fee_bnb} BNB" + (f"+{fee_other}" if fee_other else "")
        print(f"{tag:<18}{'✅ ' + (d.get('status') or ''):<8}"
              f"{str(d.get('orderId', '-')):>14}{qty:>14.6f}{quote:>18.4f}"
              f"{(d.get('avgPrice') or '-'):>14}{fee_str:>22}")
    print("-" * 100)
    print(f"{'合计':<18}{ok}/{len(selected_servers)}{'':<2}{'':>14}"
          f"{total_qty:>14.6f}{total_quote:>18.4f}{'':>14}{total_fee:>20.6f} BNB")
    print(sep)


# ============================== 流程编排 ==============================

PER_STRATEGY_ACTIONS = {
    "1": ("启动策略", start_strategy),
    "2": ("停止策略", stop_strategy),
    "3": ("查看运行状态", view_status),
    "4": ("查看日志（最后50行）", view_logs),
    "5": ("查看配置", view_config),
    "6": ("修改配置", update_config),
}


def sync_main(selected_servers):
    print("\n当前本地配置:")
    for k, spec in STRATEGIES.items():
        lp = PROJECT_DIR / spec.config
        if not lp.exists():
            continue
        cfg = load_yaml(lp)
        sym = get_nested(cfg, spec.fields.get("symbol"))
        vol = get_nested(cfg, spec.fields.get("max_volume"))
        print(f"  {spec.desc}: symbol={sym} max_volume={vol}")

    print("\n选择同步策略:")
    for k, spec in STRATEGIES.items():
        print(f"  {k}. {spec.desc} ({spec.name})")
    print("  all. 全部")
    inp = input("选择 (逗号分隔 / all): ").strip()
    if inp == "all":
        specs = list(STRATEGIES.values())
    else:
        specs = [STRATEGIES[k.strip()] for k in inp.split(",") if k.strip() in STRATEGIES]
    if not specs:
        print("❌ 无效选择")
        return

    password = get_password()
    ok = sum(1 for srv in selected_servers if sync_server(srv, password, specs))
    banner(f"同步完成: 成功 {ok}, 失败 {len(selected_servers) - ok}")


def deploy_main(selected_servers):
    print("\n部署模式:")
    print("  1. 仅代码 (不覆盖配置和 .env)")
    print("  2. 全量部署 (代码 + 配置 + .env) [默认]")
    print("  3. 完整安装 (代码 + 配置 + .env + conda + pip)")
    choice = input("请选择 (1/2/3, 默认 2): ").strip() or "2"
    code_only = choice == "1"
    full = choice == "3"
    mode_desc = "仅代码" if code_only else ("完整安装" if full else "全量部署")
    password = get_password()
    banner(f"🚀 并行部署 {len(selected_servers)} 台服务器 [{mode_desc}]")
    t0 = time.time()
    success_list, failed_list = [], []
    with ThreadPoolExecutor(max_workers=min(10, len(selected_servers))) as ex:
        futures = {ex.submit(deploy_server, s, password, full, code_only): s for s in selected_servers}
        for fut in as_completed(futures):
            try:
                ok, srv, err = fut.result()
                (success_list if ok else failed_list).append(srv if ok else (srv, err))
            except Exception as e:
                failed_list.append((futures[fut], str(e)))
                with print_lock:
                    print(f"[{futures[fut]['id']}] ❌ 异常: {e}")
    banner(f"部署完成: 成功 {len(success_list)}, 失败 {len(failed_list)} (耗时 {time.time()-t0:.1f}s)")
    if failed_list:
        print("\n❌ 失败服务器:")
        for s, err in failed_list:
            print(f"  {s['id']}. {s['ip']} - {err}")


def per_strategy_main(action_key, selected_servers):
    _, handler = PER_STRATEGY_ACTIONS[action_key]
    spec = select_strategy()
    if not spec:
        print("❌ 无效策略")
        return
    password = get_password()
    display_mode = None
    if action_key == "5":
        display_mode = "raw" if input("\n显示模式 (1.美观 / 2.原始 YAML, 默认 1): ").strip() == "2" else "formatted"
    for srv in selected_servers:
        if action_key == "5":
            handler(srv, spec, password, display_mode)
        else:
            handler(srv, spec, password)
    banner("操作完成")


def main():
    if subprocess.run("which sshpass", shell=True, capture_output=True).returncode != 0:
        print("❌ 未安装 sshpass: brew install hudochenkov/sshpass/sshpass")
        sys.exit(1)

    banner("Binance Python 策略管理工具 v2.0")
    actions = [
        *[(k, v[0]) for k, v in PER_STRATEGY_ACTIONS.items()],
        ("7", "同步配置 (.env + 策略配置)"),
        ("8", "部署代码"),
        ("9", "查询账户余额 (资金/现货/合约)"),
        ("10", "市价买入 token"),
    ]
    print("\n可用操作:")
    for k, label in actions:
        print(f"  {k:>2}. {label}")

    valid_keys = {k for k, _ in actions}
    choice = input(f"\n选择 (1-{len(actions)}): ").strip()
    if choice not in valid_keys:
        print("❌ 无效")
        sys.exit(1)

    servers = read_servers(EXCEL_PATH)
    if not servers:
        print("❌ 未找到可用服务器")
        sys.exit(1)
    selected = select_servers(servers)
    if not selected:
        print("❌ 未选择服务器")
        sys.exit(1)

    if choice == "7":
        sync_main(selected)
    elif choice == "8":
        deploy_main(selected)
    elif choice == "9":
        query_balance_main(selected)
    elif choice == "10":
        buybnb_main(selected)
    else:
        per_strategy_main(choice, selected)


if __name__ == "__main__":
    main()
